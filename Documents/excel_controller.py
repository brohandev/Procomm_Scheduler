import datetime
import os
from openpyxl import load_workbook
from prettyprinter import pprint as pp

BASE_DIR = os.path.dirname(os.path.realpath(__file__))
INPUT_DOC = BASE_DIR + "/input.xlsx"
OUTPUT_DOC = BASE_DIR + "/output.xlsx"

"""
    Allows schedule administrator to insert lunch and tea timings during specified period number. The worksheet named 
    "REQUIREMENTS" in input.xslx contains 2 columns "LUNCH PERIOD" and "TEA PERIOD", under which the administrator 
    can insert a valid period number (between 1 and 26 inclusive). Any number outside of this range will be rejected 
    and the program will exit.
"""
def retrieve_break_times():
    workbook = load_workbook(INPUT_DOC)
    sheet = workbook["REQUIREMENTS"]

    lunch_period = None
    tea_period = None

    # iterate over the columns in the sheet
    for column in sheet.iter_cols():
        column_name = column[0].value
        for i, cell in enumerate(column):
            # skip column name
            if i == 0:
                continue

            if column_name == "LUNCH PERIOD":
                lunch_period = cell.value
            elif column_name == "TEA PERIOD":
                tea_period = cell.value

    if lunch_period < 1 or lunch_period > 26:
        print("Please enter a valid value for lunch period in input.xslx (between 1 and 26 inclusive)")
        exit()
    if tea_period < 1 or tea_period > 26:
        print("Please enter a valid value for lunch period in input.xslx (between 1 and 26 inclusive)")
        exit()

    return lunch_period, tea_period


"""
    Allows schedule administrator to insert postal codes of the companies under the appropriate headings. The worksheet 
    named "COMPANIES" in input.xslx contains 2 columns "MORE PROFITABLE COMPANIES" and "LESS PROFITABLE COMPANIES", 
    under which the administrator can insert a valid postal code. 
"""
def retrieve_company_postcodes():
    workbook = load_workbook(INPUT_DOC)
    sheet = workbook["COMPANIES"]

    profitable_companies = []
    unprofitable_companies = []

    # iterate over the columns in the sheet
    for column in sheet.iter_cols():
        column_name = column[0].value
        for i, cell in enumerate(column):
            # skip column name
            if i == 0:
                continue

            if column_name == "MORE PROFITABLE COMPANIES":
                profitable_companies.append(cell.value)
            elif column_name == "LESS PROFITABLE COMPANIES":
                unprofitable_companies.append(cell.value)

    return list(filter(lambda elem: elem is not None, profitable_companies)), \
        list(filter(lambda elem: elem is not None, unprofitable_companies))


"""
    Cleans up the output.xslx Excel sheet from previous output values
"""
def clean_excel():
    try:
        wb = load_workbook(OUTPUT_DOC)
        ws = wb['SCHEDULE']

        for row in ws['A2:AA9']:
            for cell in row:
                cell.value = None

        wb.save(OUTPUT_DOC)
        wb.close()
    except Exception as e:
        pp(str(e))


"""
    Writes the final 8 Day Schedule onto the output.xslx Excel sheet
"""
def write_to_excel(path, seconds):
    duration = str(datetime.timedelta(seconds=seconds))
    path.append(duration)

    try:
        wb = load_workbook(OUTPUT_DOC)
        ws = wb['SCHEDULE']
        ws.append(path)
        wb.save(OUTPUT_DOC)
        wb.close()
    except Exception as e:
        pp(str(e))


if __name__ == '__main__':
    profitable, unprofitable = retrieve_company_postcodes()
    lunch_period, tea_period = retrieve_break_times()

    pp(profitable)
    pp(unprofitable)

    pp(lunch_period)
    pp(tea_period)
